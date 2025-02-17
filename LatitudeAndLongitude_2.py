import requests
import openpyxl

# 配置Baidu API信息
host = "https://api.map.baidu.com"
uri = "/geocoding/v3"
ak = ""  


def get_location(address):
    params = {
        "address": address,
        "output": "json",
        "ak": ak,
        "ret_coordtype": "GCJ02"
    }
    # response = requests.get(url=host + uri, params=params)
    response = requests.get(url=host + uri, params=params, verify=False)
    if response and response.status_code == 200:
        result = response.json().get("result")
        if result:
            lng = result["location"]["lng"]
            lat = result["location"]["lat"]
            return lng, lat
    return None, None



input_file = ""  
output_file = ""  

# 打开Excel文件
wb = openpyxl.load_workbook(input_file)
ws = wb.active

# 设置计数器，限制处理10条数据
max_count = 3

# 假设地址列是第1列，数据从第2行开始处理
for row in ws.iter_rows(min_row=2432, max_col=1):  # 去掉 values_only=True，保留单元格对象
    address = row[0].value  # 获取单元格的值

    # 获取当前行号
    row_num = row[0].row  # 获取当前行号

    # 如果地址是"null"，跳过当前循环
    if address == "null":
        continue

    # 获取经纬度数据
    lng, lat = get_location(address)

    if lng is not None and lat is not None:
        # 把经纬度写入相同的行，第2列为纬度，第3列为经度
        ws[f"B{row_num}"] = lat  # 假设第2列存纬度
        ws[f"C{row_num}"] = lng  # 假设第3列存经度

    # 增加计数器
    max_count -= 1
    if max_count <= 0:
        break

# 保存所有数据到文件
wb.save(output_file)

print("完成！已保存到", output_file)
